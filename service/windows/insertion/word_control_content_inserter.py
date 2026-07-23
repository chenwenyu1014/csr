#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Word Content Control 内容插入服务（完整版）
基于Content Control的智能插入，支持：
1. 检测文件方向（Word/RTF/Excel）
2. 自动分类纵向/横向占位符
3. 在控件内插入分节符和切换页面方向
4. Excel使用COM粘贴保留格式
"""

import json
import logging
import re
import time
from dataclasses import dataclass, field
import win32com.client as win32
import pywintypes
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

# 从独立模块导入数据类（不依赖win32com，可在Linux上导入）

logger = logging.getLogger(__name__)

# COM错误码
RPC_E_CALL_REJECTED = -2147418111  # 被呼叫方拒绝接收呼叫
RPC_E_SERVERCALL_RETRYLATER = -2147417846  # 服务器繁忙，稍后重试

@dataclass
class ResourceMapping:
    """资源映射"""
    placeholder: str  # 占位符（如 {{Table_1_Start}}）
    path: str  # 资源路径
    type: str  # 资源类型（table/image/excel）
    source_file: str  # 来源文件
    description: Optional[str] = None  # 描述
    orientation: Optional[str] = None  # 🆕 纸张方向（portrait/landscape），由Windows端检测后填充


@dataclass
class ContentInsertResult:
    """内容插入结果"""
    success: bool
    message: str
    output_file: Optional[str] = None
    inserted_controls: List[str] = field(default_factory=list)
    inserted_resources: List[str] = field(default_factory=list)
    error: Optional[str] = None
    # 🆕 资源方向信息（Windows端检测后返回给Linux）
    resource_orientations: Optional[dict] = None  # {占位符: "portrait"/"landscape"}

def com_retry(func, max_retries=5, delay=0.5, *args, **kwargs):
    """
    COM 操作重试辅助函数

    处理 Word/Excel COM 操作中常见的 RPC_E_CALL_REJECTED 错误
    当 COM 服务器繁忙时自动重试

    Args:
        func: 要执行的函数或 lambda
        max_retries: 最大重试次数
        delay: 重试间隔（秒），每次重试会递增
        *args, **kwargs: 传递给 func 的参数

    Returns:
        func 的返回值

    Raises:
        最后一次重试的异常
    """
    last_error = None
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except pywintypes.com_error as e:
            error_code = e.args[0] if e.args else None
            # 检查是否是可重试的 COM 错误
            if error_code in (RPC_E_CALL_REJECTED, RPC_E_SERVERCALL_RETRYLATER):
                last_error = e
                wait_time = delay * (attempt + 1)  # 递增等待时间
                logger.warning(f"⚠️ COM操作被拒绝，{wait_time:.1f}秒后重试 ({attempt + 1}/{max_retries})")
                time.sleep(wait_time)
            else:
                # 其他 COM 错误直接抛出
                raise
        except Exception as e:
            # 非 COM 错误直接抛出
            raise

    # 所有重试都失败，抛出最后一个错误
    raise last_error

# Word常量
wdCollapseEnd = 0
wdCollapseStart = 1
wdSectionBreakNextPage = 2
wdOrientLandscape = 1
wdOrientPortrait = 0


class WordControlContentInserter:
    """
    基于Content Control的智能内容插入器

    功能：
    1. 检测Word/RTF/Excel文件的纸张方向
    2. 自动分类纵向/横向占位符
    3. 在控件内插入分节符和切换页面方向
    4. Word/RTF使用InsertFile保留格式
    5. Excel使用COM粘贴保留格式
    """

    def __init__(self):
        """初始化插入器"""
        self.word = None
        self.excel = None
        logger.info("WordControlContentInserter初始化完成")

    def _connect_word(self):
        """连接Word应用"""
        # 先清理旧的 Word 实例（如果有）
        if self.word is not None:
            try:
                self.word.Quit()
            except:
                pass
            self.word = None

        # 多线程环境需要初始化COM
        import pythoncom
        try:
            pythoncom.CoInitialize()
        except:
            pass

        max_retries = 3
        last_error = None

        for attempt in range(max_retries):
            try:
                # 直接使用标准Dispatch（EnsureDispatch在某些环境有问题）
                from utils.windows_com import safe_dispatch
                self.word = safe_dispatch("Word.Application", use_ex=False, logger=logger)
                logger.info(f"✅ 连接Word应用（标准模式，尝试{attempt + 1}/{max_retries}）")

                # 验证连接是否有效
                _ = self.word.Version

                self.word.Visible = False
                self.word.DisplayAlerts = 0
                return  # 成功连接

            except Exception as e:
                last_error = e
                logger.warning(f"⚠️ Word连接失败（尝试{attempt + 1}/{max_retries}）: {e}")

                # 清理当前实例并重试
                try:
                    if self.word:
                        self.word.Quit()
                except:
                    pass
                self.word = None

                import time
                time.sleep(1)  # 等待1秒再重试

        raise RuntimeError(f"无法连接Word应用（重试{max_retries}次后失败）: {last_error}")

    def _connect_excel(self):
        """连接Excel应用"""
        if self.excel is None:
            # Flask多线程环境需要初始化COM
            import pythoncom
            try:
                pythoncom.CoInitialize()
            except:
                pass

            try:
                # 直接使用动态Dispatch
                self.excel = win32.dynamic.Dispatch("Excel.Application")
                logger.info("✅ 连接Excel应用（动态模式）")
            except:
                # 备用：标准Dispatch
                from utils.windows_com import safe_dispatch
                self.excel = safe_dispatch("Excel.Application", use_ex=False, logger=logger)
                logger.info("✅ 连接Excel应用")

            self.excel.Visible = False
            self.excel.DisplayAlerts = False

    def _cleanup(self):
        """清理资源"""
        # 清理Excel
        try:
            if self.excel:
                self.excel.Quit()
        except Exception as e:
            logger.warning(f"Excel清理异常: {e}")
        finally:
            self.excel = None

        # 清理Word
        try:
            if self.word:
                self.word.Quit()
        except Exception as e:
            logger.warning(f"Word清理异常: {e}")
        finally:
            self.word = None

        # 🆕 强制垃圾回收，确保 COM 对象被释放
        try:
            import gc
            gc.collect()
        except:
            pass

        # 释放COM资源
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except:
            pass


    @staticmethod
    def _is_table_json(text: str) -> bool:
        """
        检测是否为表格 JSON（统一格式）

        预期格式：
        {
            "table": {
                "total_rows": 行数,
                "total_cols": 列数,
                "cells": [{"r": 行号, "c": 列号, "cs": 跨列数, "rs": 跨行数, "text": "内容"}]
            }
        }

        Args:
            text: 待检测的字符串

        Returns:
            是否为表格 JSON 格式
        """
        if not isinstance(text, str) or not text.strip():
            return False

        try:
            data = json.loads(text.strip())
            if not isinstance(data, dict):
                return False
            if "table" not in data:
                return False
            t = data["table"]
            if not isinstance(t, dict):
                return False
            if "total_rows" in t and "total_cols" in t and "cells" in t:
                return True
            return False
        except (json.JSONDecodeError, TypeError, ValueError):
            return False

    @staticmethod
    def _parse_table_json(text: str) -> dict:
        """
        解析表格 JSON，返回 {"table": {...}} 结构

        Args:
            text: 表格 JSON 字符串

        Returns:
            {"table": {"total_rows": N, "total_cols": N, "cells": [...]}}
        """
        data = json.loads(text.strip())
        return {"table": data["table"]}

    def _insert_table_to_control(
        self,
        doc,
        control,
        table_data: dict,
        font_info: Tuple[str, str, int]
    ) -> bool:
        """
        插入表格（统一格式：cells 坐标式，支持跨列/跨行合并）

        Args:
            doc: Word 文档对象
            control: Content Control 对象
            table_data: {"table": {"total_rows": N, "total_cols": N, "cells": [...]}}
            font_info: (font_ascii, font_fareast, font_size) 元组

        Returns:
            是否插入成功
        """
        from collections import defaultdict

        table_info = table_data["table"]
        total_rows = table_info["total_rows"]
        total_cols = table_info["total_cols"]
        cells = table_info["cells"]
        font_ascii, font_fareast, font_size = font_info

        logger.info(f"  创建表格: {total_rows}行 x {total_cols}列, 共{len(cells)}个单元格")

        try:
            # 1. 清空控件内容
            try:
                control.Range.Delete()
            except Exception:
                pass

            # 2. 创建表格
            table = com_retry(
                lambda: doc.Tables.Add(
                    Range=control.Range,
                    NumRows=total_rows,
                    NumColumns=total_cols
                ),
                max_retries=3,
                delay=0.3
            )

            if table is None:
                logger.error("  ❌ Tables.Add 返回 None")
                return False

            # 3. 填充所有单元格内容（不合并）
            for cell_def in cells:
                r = cell_def["r"]
                c = cell_def["c"]
                text = cell_def.get("text", "")
                try:
                    com_cell = table.Cell(r + 1, c + 1)
                    com_cell.Range.Text = text if text else ""
                    # 设置字体
                    com_cell.Range.Font.NameAscii = font_ascii
                    com_cell.Range.Font.NameFarEast = font_fareast
                    com_cell.Range.Font.Size = font_size
                    com_cell.Range.Font.Bold = False
                except Exception as e:
                    logger.warning(f"  填充单元格 ({r},{c}) 失败: {e}")

            # 4. 合并单元格
            #    策略：先处理纯跨列和纯跨行，最后处理 both（拆成每行跨列+跨行两步）
            #    这样避免 both 的合并影响纯跨列/跨行的坐标
            only_cs = [d for d in cells if d.get("cs", 1) > 1 and d.get("rs", 1) <= 1]
            only_rs = [d for d in cells if d.get("cs", 1) <= 1 and d.get("rs", 1) > 1]
            both = [d for d in cells if d.get("cs", 1) > 1 and d.get("rs", 1) > 1]

            # 4a. 纯跨列合并（从右往左）
            for cell_def in sorted(only_cs, key=lambda x: (x["r"], -x["c"])):
                r, c = cell_def["r"], cell_def["c"]
                cs = cell_def["cs"]
                try:
                    table.Cell(r + 1, c + 1).Merge(table.Cell(r + 1, c + cs))
                except Exception as e:
                    logger.warning(f"  跨列合并失败 ({r},{c},cs={cs}): {e}")

            # 4b. 纯跨行合并（从下往上）
            for cell_def in sorted(only_rs, key=lambda x: (-x["r"], x["c"])):
                r, c = cell_def["r"], cell_def["c"]
                rs = cell_def["rs"]
                try:
                    table.Cell(r + 1, c + 1).Merge(table.Cell(r + rs, c + 1))
                except Exception as e:
                    logger.warning(f"  跨行合并失败 ({r},{c},rs={rs}): {e}")

            # 4c. 最后处理同时跨行+跨列：拆成两步
            for cell_def in sorted(both, key=lambda x: (-x["r"], -x["c"])):
                r, c = cell_def["r"], cell_def["c"]
                cs = cell_def["cs"]
                rs = cell_def["rs"]

                # Step 1: 每行做跨列合并（从下往上处理）
                for row_idx in range(r + rs - 1, r - 1, -1):
                    try:
                        table.Cell(row_idx + 1, c + 1).Merge(table.Cell(row_idx + 1, c + cs))
                    except Exception as e:
                        logger.warning(f"  行{row_idx} 跨列合并失败: {e}")

                # Step 2: 跨行合并
                try:
                    table.Cell(r + 1, c + 1).Merge(table.Cell(r + rs, c + 1))
                except Exception as e:
                    logger.warning(f"  跨行合并失败 ({r},{c},rs={rs}): {e}")

            # 5. 自动调整列宽（根据内容）
            try:
                table.AutoFitBehavior(1)  # wdAutoFitContent = 1
            except Exception as e:
                logger.warning(f"  自动调整列宽失败: {e}")

            for col_idx in range(1, total_cols + 1):
                for row_idx in range(1, total_rows + 1):
                    try:
                        com_cell = table.Cell(row_idx, col_idx)
                        com_cell.Range.ParagraphFormat.SpaceAfter = 0
                        com_cell.Range.ParagraphFormat.SpaceBefore = 0
                        com_cell.Range.ParagraphFormat.LineSpacingRule = 0
                    except Exception:
                        # 被合并掉的单元格访问会报错，直接跳过
                        pass

            # 6. 设置表格边框
            try:
                borders = table.Borders
                borders.Enable = 1  # 开启所有边框
                borders.InsideLineStyle = 1  # wdLineStyleSingle
                borders.OutsideLineStyle = 1  # wdLineStyleSingle
                borders.InsideLineWidth = 4   # 1 磅
                borders.OutsideLineWidth = 4  # 1 磅
            except Exception as e:
                logger.debug(f"  设置表格边框失败: {e}")

            logger.info(f"  表格插入成功")
            return True

        except Exception as e:
            import traceback
            traceback.print_exc()
            logger.error(f"  表格插入失败: {e}")
            return False

    @staticmethod
    def _clean_text_for_word(text: str) -> str:
        """
        清理文本中可能导致 Word COM 报错的字符

        Word COM 的 InsertAfter / Range.Text 对以下字符零容忍：
        - 不间断空格 \u00A0
        - 零宽字符 \u200B \u200C \u200D \uFEFF
        - 软连字符 \u00AD
        - 行尾不可见空格（Markdown two-space line break）
        - 其他 Unicode 特殊空白和控制字符
        """
        if not text:
            return ""

        import re

        # 1. 统一换行符
        text = text.replace('\r\n', '\r').replace('\n', '\r')

        # 2. 替换各种特殊空白为普通空格
        special_spaces = [
            '\u00A0',  # 不间断空格 (Non-Breaking Space)
            '\u2002',  # En Space
            '\u2003',  # Em Space
            '\u2004',  # Three-Per-Em Space
            '\u2005',  # Four-Per-Em Space
            '\u2006',  # Six-Per-Em Space
            '\u2007',  # Figure Space
            '\u2008',  # Punctuation Space
            '\u2009',  # Thin Space
            '\u200A',  # Hair Space
            '\u202F',  # Narrow No-Break Space
            '\u205F',  # Medium Mathematical Space
            '\u3000',  # 全角空格 (Ideographic Space)
        ]
        for sp in special_spaces:
            text = text.replace(sp, ' ')

        # 3. 删除零宽字符和不可见字符
        invisible_chars = [
            '\u200B',  # 零宽空格 (Zero Width Space)
            '\u200C',  # 零宽不连字 (Zero Width Non-Joiner)
            '\u200D',  # 零宽连字 (Zero Width Joiner)
            '\u200E',  # 从左到右标记
            '\u200F',  # 从右到左标记
            '\uFEFF',  # BOM / 零宽不间断空格
            '\u00AD',  # 软连字符 (Soft Hyphen)
            '\u2028',  # 行分隔符 (Line Separator)
            '\u2029',  # 段分隔符 (Paragraph Separator)
        ]
        for ch in invisible_chars:
            text = text.replace(ch, '')

        # 4. 清理 Markdown 行尾两个空格（LLM 常见输出，会产生不可见尾部空格）
        #    "内容  \n" -> "内容\n"
        text = re.sub(r'  +\r', '\r', text)
        text = re.sub(r' +\r', '\r', text)

        # 5. 逐字符过滤残余控制字符
        #    保留：普通可打印字符、\n（换行）、\t（制表符）
        #    替换：其他 ASCII 控制字符（\x00-\x1F 中除 \t \n 外）
        cleaned = []
        for ch in text:
            code = ord(ch)
            if ch in ('\r', '\t'):
                cleaned.append(ch)
            elif code < 0x20:
                # 其他控制字符用空格替代（不直接删除，避免词语粘连）
                cleaned.append(' ')
            elif code == 0xFFFD:
                # Unicode 替换字符
                cleaned.append(' ')
            else:
                cleaned.append(ch)
        text = ''.join(cleaned)

        # 6. 压缩连续空行（超过2个换行压缩为2个）
        text = re.sub(r'\r{3,}', '\r\r', text)

        # 7. 去掉首尾空白
        text = text.strip()

        return text

    def detect_file_orientation(self, file_path: str) -> str:
        """
        检测文件的纸张方向

        Args:
            file_path: 文件路径

        Returns:
            "landscape" 或 "portrait"
        """
        file_ext = Path(file_path).suffix.lower()

        try:
            # Word文件：不需要检测方向（InsertFile会保留原格式）
            if file_ext in ['.docx', '.doc']:
                logger.info(f"  Word文件跳过方向检测: {Path(file_path).name}")
                return "portrait"

            # RTF文件：通过读取文件内容判断方向（不需要打开 Word）
            # RTF 格式中：
            # 1. \landscape 控制字表示横向
            # 2. 如果没有 \landscape，通过 \paperw 和 \paperh 判断（宽度 > 高度 = 横向）
            elif file_ext == '.rtf':
                logger.info(f"  🔍 检测RTF方向（文件内容扫描）: {Path(file_path).name}")
                try:
                    landscape_found = False
                    paper_width = None
                    paper_height = None
                    max_lines = 10  # 只读取前10行

                    # 逐行读取文件的前几行
                    with open(file_path, 'r', encoding='latin1', errors='ignore') as file:
                        for i, line in enumerate(file):
                            if i >= max_lines:  # 只读取前max_lines行
                                break

                            # 检查是否有 \landscape
                            if '\\landscape' in line:
                                landscape_found = True
                                break  # 找到横向后，直接跳出循环

                            # 查找纸张宽度和高度
                            width_match = re.search(r'\\paperw(\d+)', line)
                            height_match = re.search(r'\\paperh(\d+)', line)

                            if width_match:
                                paper_width = int(width_match.group(1))
                            if height_match:
                                paper_height = int(height_match.group(1))

                            # 如果同时找到了宽度和高度，可以提前结束循环
                            if paper_width and paper_height:
                                break

                    # 判断纸张方向
                    if landscape_found:
                        logger.info(f"    ✅ RTF方向: landscape (检测到 \\landscape)")
                        return "landscape"

                    if paper_width is not None and paper_height is not None:
                        if paper_width > paper_height:
                            logger.info(f"    ✅ RTF方向: landscape (宽度{paper_width} > 高度{paper_height})")
                            return "landscape"
                        else:
                            logger.info(f"    ✅ RTF方向: portrait (宽度{paper_width} ≤ 高度{paper_height})")
                            return "portrait"

                    # 如果都没有找到，默认横向（RTF 通常是数据表格）
                    logger.info(f"    ✅ RTF方向: landscape (默认，RTF通常是数据表格)")
                    return "landscape"

                except Exception as e:
                    logger.warning(f"    ⚠️ RTF内容读取失败: {e}，默认横向")
                    return "landscape"

            # Excel文件：通过表格形状判断
            elif file_ext in ['.xlsx', '.xls']:
                logger.info(f"  🔍 开始检测Excel方向: {Path(file_path).name}")
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb.active

                    max_row = ws.max_row
                    max_col = ws.max_column
                    wb.close()

                    logger.info(f"    Excel尺寸: {max_row}行 × {max_col}列")

                    # 🆕 综合判断逻辑
                    # 小表格（行数<10 且列数<10）默认纵向，除非明显宽表
                    is_small_table = max_row < 10 and max_col < 10
                    if is_small_table:
                        # 小表格：只有列数明显多于行数（2 倍）才横向
                        if max_col > max_row * 2:
                            logger.info(f"    ✅ Excel 方向：landscape (小表格，列数>{max_row}×2)")
                            return "landscape"
                        else:
                            logger.info(f"    ✅ Excel 方向：portrait (小表格，适合纵向)")
                            return "portrait"
                    else:
                        # 大表格：使用比例判断
                        if max_col > max_row * 1.5 and max_col >= 8:
                            logger.info(f"    ✅ Excel 方向：landscape (大表格，明显宽表)")
                            return "landscape"
                        else:
                            logger.info(f"    ✅ Excel 方向：portrait (大表格，适合纵向)")
                            return "portrait"

                except Exception as e:
                    logger.warning(f"    ⚠️ Excel方向检测失败: {e}，默认纵向")
                    return "portrait"  # 默认纵向

            # 其他文件：默认纵向
            else:
                return "portrait"

        except Exception as e:
            logger.error(f"检测失败 {Path(file_path).name}: {e}")
            return "portrait"

    def classify_placeholders_by_orientation(
            self,
            placeholders: List[str],
            resource_mappings: Dict[str, ResourceMapping]
    ) -> Tuple[List[str], List[str]]:
        """
        根据文件纸张方向分类占位符

        🆕 优化：将检测结果缓存到 mapping.orientation，便于后续使用

        Args:
            placeholders: 占位符列表
            resource_mappings: 资源映射

        Returns:
            (portrait_placeholders, landscape_placeholders)
            纵向占位符在前，横向占位符在后，便于插入时只用一对分节符
        """
        portrait = []
        landscape = []

        for placeholder in placeholders:
            if placeholder not in resource_mappings:
                logger.warning(f"占位符无映射: {placeholder}")
                portrait.append(placeholder)
                continue

            mapping = resource_mappings[placeholder]
            file_path = mapping.path

            if not Path(file_path).exists():
                logger.warning(f"文件不存在: {file_path}")
                portrait.append(placeholder)
                continue

            # 检测方向
            orientation = self.detect_file_orientation(file_path)

            # 🆕 缓存方向信息到 mapping（如果支持动态属性）
            try:
                mapping.orientation = orientation
            except:
                pass

            if orientation == "landscape":
                landscape.append(placeholder)
                logger.info(f"  📐 横向: {placeholder} <- {Path(file_path).name}")
            else:
                portrait.append(placeholder)
                logger.info(f"  📄 纵向: {placeholder} <- {Path(file_path).name}")

        # 🆕 日志汇总
        logger.info(f"  📊 方向分类完成: 纵向{len(portrait)}个, 横向{len(landscape)}个")

        return portrait, landscape

    def insert_to_control(
        self,
        doc,
        control,
        generated_text: str,
        portrait_placeholders: List[str],
        landscape_placeholders: List[str],
        font_tuple: Optional[Tuple[str, str, int]] = None
    ):
        """
        在控件内插入内容

        核心策略（参考原始设计）：
        - 纵向内容（包含纵向占位符）直接插入控件
        - 所有横向占位符集中到一起，用一对分节符包裹（纵向→横向→纵向）
        - 这样只有2个分节符，而不是每个横向文件各2个

        Args:
            doc: Word文档对象
            control: Content Control对象
            generated_text: 生成的文本
            portrait_placeholders: 纵向占位符列表
            landscape_placeholders: 横向占位符列表
        """
        # 使用传入的字体或从控件获取
        if font_tuple:
            font_ascii, font_fareast, font_size = font_tuple
        else:
            # 自动获取控件本身的格式（向后兼容）
            control_font_ascii = None
            control_font_fareast = None
            control_font_size = None
            try:
                if control.Range.Font:
                    control_font_ascii = control.Range.Font.NameAscii
                    control_font_fareast = control.Range.Font.NameFarEast
                    control_font_size = control.Range.Font.Size
            except Exception:
                pass

            font_ascii = control_font_ascii if control_font_ascii else "Times New Roman"
            # 因为 Times New Roman 等西文字体不能作为 NameFarEast 的值，使用默认中文字体（宋体）
            if control_font_fareast and control_font_fareast != "Times New Roman":
                font_fareast = control_font_fareast
            else:
                font_fareast = "宋体"
                logger.warning(f"  ⚠️ 控件中文字体不能为Times New Roman，使用默认中文字体: {font_fareast}")
            font_size = control_font_size if control_font_size else 12
        logger.info(f"   控件字体：西文={font_ascii}, 中文={font_fareast}, 字号={font_size}")

        # 重新排列内容：从文本中移除横向占位符（它们会被集中到横向节）
        portrait_content = generated_text
        landscape_content = ""

        for p in landscape_placeholders:
            portrait_content = portrait_content.replace(p, "")  # 移除横向占位符
            landscape_content += f"{p}\n"  # 收集横向占位符

        # 清理可能导致 COM 错误的字符
        portrait_content = self._clean_text_for_word(portrait_content)
        control.Range.Delete()
        control.Range.Style = "正文"

        # 尝试插入文本，有备用方案处理 COM 错误
        try:
            control.Range.InsertAfter(portrait_content)
            logger.info(f"  ✅ 插入纵向内容 ({len(portrait_content)} 字符)，手动应用控件样式")
        except pywintypes.com_error as insert_err:
            # InsertAfter 失败，尝试备用方案
            logger.warning(f"  ⚠️ InsertAfter 失败，尝试分段插入: {insert_err}")
            try:
                # 分段处理：将内容按段落分开，逐段插入
                paragraphs = portrait_content.split('\r')
                for para_idx, para in enumerate(paragraphs):
                    if para_idx > 0:
                        try:
                            control.Range.InsertAfter('\v')  # 垂直制表符
                        except:
                            control.Range.InsertAfter(' ')
                            logger.warning("  ⚠️ 无法插入段落标记，用空格代替")
                    if para.strip():
                        control.Range.InsertAfter(para)
                logger.info(f"  ✅ 通过分段 InsertAfter 插入成功 ({len(portrait_content)} 字符)")
            except Exception as fallback_err:
                logger.error(f"  ❌ 分段插入也失败: {fallback_err}")
                raise

        # 应用控件本身的字体格式(同时设置西文和中文)
        control.Range.Font.NameAscii = font_ascii
        control.Range.Font.NameFarEast = font_fareast
        control.Range.Font.Size = font_size
        if landscape_content:
            # 在控件末尾插入第一个分节符（切换到横向）
            rng = control.Range
            rng.Collapse(wdCollapseEnd)
            rng.InsertBreak(Type=wdSectionBreakNextPage)
            logger.info("  ✅ 插入分节符（纵向→横向）")

            # 设置新节为横向
            new_section = doc.Sections(doc.Sections.Count)
            new_section.PageSetup.Orientation = wdOrientLandscape
            logger.info("  ✅ 设置横向页面")

            # 插入所有横向占位符（集中在一起）
            newrng = doc.Range(rng.Start, rng.Start)
            newrng.InsertAfter(landscape_content)
            newrng.Font.NameAscii = font_ascii
            newrng.Font.NameFarEast = font_fareast
            newrng.Font.Size = font_size
            logger.info(f"  ✅ 插入 {len(landscape_placeholders)} 个横向占位符")

            # 在横向区域末尾插入第二个分节符（恢复纵向）
            newrng.Collapse(wdCollapseEnd)
            newrng.InsertBreak(Type=wdSectionBreakNextPage)
            logger.info("  ✅ 插入分节符（横向→纵向）")

            # 恢复纵向
            final_section = doc.Sections(doc.Sections.Count)
            final_section.PageSetup.Orientation = wdOrientPortrait
            logger.info("  ✅ 恢复纵向页面")

    def _get_range_by_scanning(self, ws, max_scan_rows=100, max_scan_cols=100):
        """
        通过扫描单元格获取有效范围
        """
        logger.info("扫描单元格获取有效范围...")
        # 确定第一个和最后一个有内容的单元格
        first_row = None
        first_col = None
        last_row = None
        last_col = None

        # 扫描前max_scan_rows行和max_scan_cols列
        for row in range(1, max_scan_rows + 1):
            row_has_content = False

            for col in range(1, max_scan_cols + 1):
                try:
                    cell = ws.Cells(row, col)
                    # 检查单元格是否有内容（值或公式）
                    has_value = cell.Value is not None
                    has_formula = False

                    try:
                        formula = cell.Formula
                        has_formula = formula is not None and formula != ''
                    except:
                        pass

                    # 如果有内容
                    if has_value or has_formula:
                        row_has_content = True

                        # 更新边界
                        if first_row is None:
                            first_row = row
                        if first_col is None or col < first_col:
                            first_col = col

                        last_row = row
                        if last_col is None or col > last_col:
                            last_col = col

                except Exception as e:
                    continue

            # 如果连续5行都没有内容，提前结束扫描
            if not row_has_content:
                if row > 50:  # 至少扫描100行
                    empty_count = 0
                    # 检查后面几行
                    for next_row in range(row + 1, min(row + 5, max_scan_rows + 1)):
                        next_row_empty = True
                        for col in range(1, max_scan_cols + 1):
                            try:
                                cell = ws.Cells(next_row, col)
                                if cell.Value is not None:
                                    next_row_empty = False
                                    break
                            except:
                                pass
                        if next_row_empty:
                            empty_count += 1
                        else:
                            empty_count = 0
                            break

                    if empty_count >= 5:  # 连续5行空行，认为已经结束
                        logger.info(f"检测到连续空行，在第{row}行停止扫描")
                        break
        # 设置默认值
        if first_row is None:
            first_row = 1
        if first_col is None:
            first_col = 1
        if last_row is None:
            last_row = first_row
        if last_col is None:
            last_col = first_col
        # 确保范围有效
        if last_row < first_row:
            last_row = first_row
        if last_col < first_col:
            last_col = first_col
        # 创建范围
        true_range = ws.Range(
            ws.Cells(first_row, first_col),
            ws.Cells(last_row, last_col)
        )

        logger.info(f"扫描结果: {true_range.Address} ({last_row - first_row + 1}行×{last_col - first_col + 1}列)")

        return true_range

    def insert_to_template(
        self,
        template_file: str,
        generation_results: List[Dict[str, Any]],
        output_file: str
    ) -> ContentInsertResult:
        """
        将生成结果插入到Word模板

        Args:
            template_file: Word模板文件路径
            generation_results: 生成结果列表 [
                {
                    "paragraph_id": "study_population",
                    "control_title": "study_population",
                    "generated_content": "文本内容\n{{Table_1}}\n更多文本",
                    "status": "success",
                    "resource_mappings": {
                        "{{Table_1}}": { "path": "...", "type": "table", ... }
                    }
                }
            ]
            output_file: 输出文件路径

        Returns:
            ContentInsertResult
        """
        try:
            logger.info("=" * 70)
            logger.info("开始插入内容到Word模板")
            logger.info("=" * 70)

            # 连接Word
            self._connect_word()

            # import pythoncom
            # pythoncom.CoInitialize()
            # import win32com as win32com
            # word = win32com.client.Dispatch("Word.Application")
            # word.Visible = False  # 显示Word窗口
            # word.DisplayAlerts = 0


            # 验证模板文件存在
            template_path = Path(template_file).absolute()
            if not template_path.exists():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")

            # 打开模板
            logger.info(f"📄 打开模板文件: {template_path}")

            try:
                # 使用 com_retry 处理 Word 繁忙的情况
                doc = com_retry(
                    lambda: self.word.Documents.Open(
                        str(template_path),
                        ConfirmConversions=False,
                        ReadOnly=False,
                        AddToRecentFiles=False
                    ),
                    max_retries=5,
                    delay=0.5
                )

                # doc = word.Documents.Open(
                #         str(template_path),
                #         ConfirmConversions=False,
                #         ReadOnly=False,
                #         AddToRecentFiles=False
                #     )

                # 验证文档对象
                if doc is None:
                    raise RuntimeError(f"Word打开文档失败，返回None")

                # 使用 com_retry 访问 doc.Name，防止 COM 繁忙错误
                doc_name = com_retry(lambda: doc.Name, max_retries=5, delay=0.3)
                logger.info(f"✅ 打开模板: {doc_name}")

                doc_name = doc.Name

            except AttributeError as e:
                # 如果仍然出错，说明 EnsureDispatch 也没有生效
                logger.error(f"❌ doc 对象类型: {type(doc)}")
                raise RuntimeError(f"Word返回的不是有效的文档对象. 错误: {e}")
            except pywintypes.com_error as e:
                error_code = e.args[0] if e.args else None
                logger.error(f"❌ Word COM 错误 (代码: {error_code}): {e}")
                raise RuntimeError(f"Word打开文档失败(COM错误): {e}")
            except Exception as e:
                raise RuntimeError(f"Word打开文档失败: {e}")

            inserted_controls = []
            inserted_resources = []
            resource_orientations = {}  # 🆕 收集方向信息
            control_fonts = {}  # 🆕 保存控件字体格式 {placeholder: (font_ascii, font_fareast, font_size)}

            # ===== 第一步：插入文本内容（包括占位符作为文本） =====
            logger.info("=" * 70)
            logger.info("第一步：插入文本内容和占位符")
            logger.info("=" * 70)

            # 处理每个段落
            for result in generation_results:
                if result.get("status") != "success":
                    continue

                control_title = result.get("control_title") or result.get("paragraph_id")
                generated_content = result.get("generated_content", "")

                logger.info(f"处理控件: {control_title}")

                # 查找控件
                cc_collection = doc.SelectContentControlsByTitle(control_title)
                if cc_collection.Count < 1:
                    logger.warning(f"  ⚠️ 未找到控件: {control_title}")
                    continue
                else:
                    logger.info(f"  找到控件: {control_title},共{cc_collection.Count}个")

                controls = cc_collection

                #  检测是否为表格 JSON：如果是，直接创建表格，跳过占位符逻辑
                if self._is_table_json(generated_content):
                    logger.info(f"  📊 检测到表格 JSON，将创建 Word 表格")
                    table_data = self._parse_table_json(generated_content)

                    # 获取控件字体格式
                    ctrl_font_ascii = 'Times New Roman'
                    ctrl_font_fareast = '宋体'
                    ctrl_font_size = 12
                    try:
                        if cc_collection.Count > 0:
                            first_control = cc_collection.Item(1)
                            if first_control.Range.Font:
                                ctrl_font_ascii = first_control.Range.Font.NameAscii or 'Times New Roman'
                                fareast = first_control.Range.Font.NameFarEast
                                if fareast and fareast != 'Times New Roman':
                                    ctrl_font_fareast = fareast
                                ctrl_font_size = first_control.Range.Font.Size or 12
                    except Exception as font_err:
                        logger.warning(f'  获取控件字体失败：{font_err}，使用默认值')

                    font_info = (ctrl_font_ascii, ctrl_font_fareast, ctrl_font_size)

                    # 插入表格
                    for i in range(cc_collection.Count):
                        try:
                            control = cc_collection.Item(i + 1)
                        except Exception as e:
                            logger.warning(f"  ⚠️ 控件 {control_title} 第{i+1}个无法访问（可能被同名外层控件覆盖），跳过: {e}")
                            continue
                        success = self._insert_table_to_control(
                            doc, control, table_data, font_info
                        )
                        if success:
                            inserted_controls.append(control_title)
                        else:
                            logger.warning(f"  ⚠️ 表格插入失败，将降级为文本插入")
                            # 降级：将 JSON 作为普通文本插入
                            cleaned_content = self._clean_text_for_word(generated_content)
                            try:
                                control.Range.Text = cleaned_content
                            except Exception:
                                pass
                    continue  # 跳过后面的占位符提取和文本插入逻辑

                # 原有逻辑：提取占位符
                placeholders = re.findall(r'\{\{[^}]+\}\}', generated_content)
                if len(placeholders):
                    logger.info(f"  发现 {len(placeholders)} 个占位符")

                # 从该段落获取 resource_mappings
                paragraph_mappings = result.get("resource_mappings", {})

                # 过滤有效的占位符（使用该段落的映射）
                valid_placeholders = [p for p in placeholders if p in paragraph_mappings]

                if not valid_placeholders:
                    # 没有占位符，直接插入文本
                    # 🆕 先获取控件字体格式
                    ctrl_font_ascii = 'Times New Roman'
                    ctrl_font_fareast = '宋体'
                    ctrl_font_size = 12
                    try:
                        if cc_collection.Count > 0:
                            first_control = cc_collection.Item(1)
                            if first_control.Range.Font:
                                ctrl_font_ascii = first_control.Range.Font.NameAscii or 'Times New Roman'
                                ctrl_font_fareast = first_control.Range.Font.NameFarEast or '宋体'
                                ctrl_font_size = first_control.Range.Font.Size or 12
                    except Exception as font_err:
                        logger.warning(f'  获取控件字体失败：{font_err}，使用默认值')

                    for i in range(cc_collection.Count):
                        try:
                            control = cc_collection.Item(i + 1)
                            # 清理内容中可能导致 COM 错误的字符
                            cleaned_content = self._clean_text_for_word(generated_content)

                            # 设置文本内容
                            try:
                                control.Range.Text = cleaned_content
                                logger.info(f"  ✅ 文本插入成功 ({len(cleaned_content)} 字符)")
                            except pywintypes.com_error as text_err:
                                # 如果直接设置 Text 失败，使用分段插入备用方案
                                logger.warning(f"  ⚠️ 设置 Text 属性失败，尝试分段插入: {text_err}")
                                try:
                                    # 先删除内容，再分段插入
                                    control.Range.Delete()

                                    # 分段处理：将内容按段落分开，逐段插入
                                    paragraphs = cleaned_content.split('\r')
                                    for para_idx, para in enumerate(paragraphs):
                                        if para_idx > 0:
                                            # 插入段落标记（换行）- 尝试多种方式
                                            try:
                                                # 使用 垂直制表符
                                                control.Range.InsertAfter('\v')
                                            except:
                                                # 最后用空格代替
                                                control.Range.InsertAfter(' ')
                                                logger.warning("  ⚠️ 无法插入段落标记，用空格代替")
                                        if para.strip():
                                            control.Range.InsertAfter(para)

                                    # 🆕 分段插入后应用控件字体格式
                                    control.Range.Font.NameAscii = ctrl_font_ascii
                                    control.Range.Font.NameFarEast = ctrl_font_fareast
                                    control.Range.Font.Size = ctrl_font_size

                                    logger.info(f"  ✅ 通过分段 InsertAfter 插入成功 ({len(cleaned_content)} 字符)")

                                except Exception as insert_err:
                                    logger.error(f"  ❌ 分段插入也失败: {insert_err}")
                                    raise

                            inserted_controls.append(control_title)
                        except Exception as e:
                            # import traceback
                            # traceback.print_exc()
                            logger.error(f"  ❌ 文本插入失败: {e}")
                            logger.error(f"  内容长度: {len(generated_content)} 字符")
                            logger.error(f"  内容前200字符: {generated_content[:200]}")
                            continue  # 跳过该控件，不中断整篇文档（防护同名嵌套等异常）
                    continue

                # 保存占位符对应的字体格式
                ctrl_font_ascii = 'Times New Roman'
                ctrl_font_fareast = '宋体'
                ctrl_font_size = 12
                try:
                    if controls.Count > 0:
                        first_control = controls.Item(1)
                        if first_control.Range.Font:
                            ctrl_font_ascii = first_control.Range.Font.NameAscii or 'Times New Roman'
                            # Times New Roman 不是有效的中文字体，需要检查并使用默认值
                            fareast = first_control.Range.Font.NameFarEast
                            if fareast and fareast != 'Times New Roman':
                                ctrl_font_fareast = fareast
                            # 否则保持默认值 '宋体'
                            ctrl_font_size = first_control.Range.Font.Size or 12
                except Exception as font_err:
                    logger.warning(f'  获取控件字体失败：{font_err}，使用默认值')

                # 保存占位符对应的字体格式（使用段落ID+占位符作为key避免冲突）
                for p in valid_placeholders:
                    control_fonts[(control_title, p)] = (ctrl_font_ascii, ctrl_font_fareast, ctrl_font_size)

                # 🆕 将段落映射转换为 ResourceMapping 对象（用于方向检测）
                paragraph_resource_mappings = {}
                for placeholder, mapping_info in paragraph_mappings.items():
                    if isinstance(mapping_info, ResourceMapping):
                        paragraph_resource_mappings[placeholder] = mapping_info
                    elif isinstance(mapping_info, dict):
                        # 从字典转换为 ResourceMapping
                        paragraph_resource_mappings[placeholder] = ResourceMapping(
                            placeholder=placeholder,
                            path=mapping_info.get("path", ""),
                            type=mapping_info.get("type", ""),
                            source_file=mapping_info.get("source_file", "")
                        )
                    else:
                        logger.warning(f"  ⚠️ 未知的映射类型: {type(mapping_info)}")

                # 检测方向并分类（使用该段落的映射）
                logger.info("  检测文件方向...")
                portrait_list, landscape_list = self.classify_placeholders_by_orientation(
                    valid_placeholders,
                    paragraph_resource_mappings
                )

                # 🆕 收集方向信息（用于返回给Linux）
                for p in portrait_list:
                    resource_orientations[p] = "portrait"
                for p in landscape_list:
                    resource_orientations[p] = "landscape"

                logger.info(f"  分类结果: 纵向{len(portrait_list)}个, 横向{len(landscape_list)}个")

                # 在控件内插入内容（占位符作为文本插入）
                for control in controls:
                    try:
                        self.insert_to_control(
                            doc,
                            control,
                            generated_content,
                            portrait_list,
                            landscape_list,
                            font_tuple=(ctrl_font_ascii, ctrl_font_fareast, ctrl_font_size)
                        )
                    except Exception as e:
                        logger.warning(f"  ⚠️ 控件 {control_title} 某个实例插入失败（可能被同名外层控件覆盖），跳过: {e}")
                        continue

                inserted_controls.append(control_title)

            # 保存并关闭文档（第一步完成）
            temp_output = str(Path(output_file).absolute())
            doc.SaveAs(temp_output)
            logger.info(f"✅ 第一步完成，保存文档: {Path(output_file).name}")

            # 关闭文档
            try:
                doc.Close(False)
                logger.info("✅ 文档已关闭")
            except Exception as close_err:
                logger.warning(f"关闭文档时出错: {close_err}")

            # 清理 COM 对象
            doc = None
            self._cleanup()

            # 🆕 等待 Word 进程完全退出，避免 RPC 连接到僵尸进程
            import time
            time.sleep(2)
            logger.info("⏳ 等待 Word 进程完全退出...")

            # ===== 第二步：重新打开文档，替换占位符为实际文件 =====
            # 🆕 采用新策略：按占位符位置遍历，反向查找所在控件
            # 判定条件：检查是否有任何段落包含 resource_mappings
            has_resource_mappings = any(
                result.get("resource_mappings") for result in generation_results
            )

            if has_resource_mappings:
                logger.info("=" * 70)
                logger.info("第二步：替换占位符为实际文件（按位置遍历）")
                logger.info("=" * 70)

                # 重新连接 Word
                max_connect_retries = 3
                for connect_attempt in range(max_connect_retries):
                    try:
                        self._connect_word()
                        doc = self.word.Documents.Open(temp_output)
                        logger.info(f"✅ 重新打开文档: {Path(output_file).name}")
                        break
                    except Exception as connect_err:
                        logger.warning(f"⚠️ 第二步连接 Word 失败 (尝试 {connect_attempt + 1}/{max_connect_retries}): {connect_err}")
                        if connect_attempt < max_connect_retries - 1:
                            self._cleanup()
                            import time
                            time.sleep(3)
                        else:
                            logger.error("❌ 无法重新连接 Word，跳过占位符替换")
                            return ContentInsertResult(
                                success=True,
                                message="文本内容已插入，但占位符替换失败",
                                output_file=output_file,
                                inserted_controls=inserted_controls,
                                inserted_resources=[],
                                resource_orientations=resource_orientations,
                                error=f"第二步失败: {connect_err}"
                            )

                # 🆕 构建控件标题到段落信息的映射
                control_to_paragraph = {}
                for result in generation_results:
                    control_title = result.get("paragraph_id")
                    paragraph_mappings = result.get("resource_mappings", {})
                    if paragraph_mappings:
                        control_to_paragraph[control_title] = {
                            "paragraph_id": result.get("paragraph_id"),
                            "mappings": paragraph_mappings
                        }

                # 🆕 按文档顺序遍历所有占位符
                logger.info("开始按位置遍历占位符...")
                replace_errors = []
                search_start = 0

                while search_start < doc.Content.End - 4:
                    # 搜索 "{{" 作为占位符起始标记
                    find_range = doc.Range(search_start, doc.Content.End)
                    find_range.Find.ClearFormatting()
                    find_range.Find.Text = "{{"
                    find_range.Find.Wrap = 0  # wdFindStop：不回绕，避免替换失败时反复命中同一占位符导致死循环

                    if not find_range.Find.Execute():
                        break

                    placeholder_start = find_range.Start

                    # 提取完整占位符 {{...}}
                    placeholder_end = min(placeholder_start + 50, doc.Content.End)
                    try:
                        text_range = doc.Range(placeholder_start, placeholder_end)
                        text = text_range.Text
                        import re as regex_module
                        match = regex_module.match(r'\{\{[^}]+\}\}', text)
                        if not match:
                            search_start = placeholder_start + 2
                            continue
                        placeholder = match.group(0)
                    except Exception:
                        search_start = placeholder_start + 2
                        continue

                    logger.info(f"  找到占位符: {placeholder} (位置: {placeholder_start})")

                    # 检查占位符所在的控件
                    try:
                        check_range = doc.Range(placeholder_start, placeholder_start)
                        parent_cc = check_range.ParentContentControl

                        if parent_cc is None:
                            logger.info(f"    ⏭️ 不在控件内，跳过")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        control_title = parent_cc.Title if hasattr(parent_cc, 'Title') else None
                        if not control_title:
                            logger.warning(f"    ⚠️ 控件无标题，跳过")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        logger.info(f"    所在控件: {control_title}")

                        # 查找控件对应的段落信息
                        paragraph_info = control_to_paragraph.get(control_title)
                        if not paragraph_info:
                            logger.warning(f"    ⚠️ 未找到控件 '{control_title}' 的映射信息，跳过")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        # 获取占位符对应的文件映射
                        mapping_info = paragraph_info["mappings"].get(placeholder)
                        if not mapping_info:
                            logger.warning(f"    ⚠️ 占位符 '{placeholder}' 无映射，跳过")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        # 获取文件路径（只需要 path 属性）
                        if isinstance(mapping_info, dict):
                            file_path = mapping_info.get("path", "")
                        elif isinstance(mapping_info, ResourceMapping):
                            file_path = mapping_info.path
                        else:
                            logger.warning(f"    ⚠️ 未知的映射类型")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        # 获取字体信息
                        font_info = control_fonts.get((control_title, placeholder))
                        if font_info:
                            font_ascii, font_fareast, font_size = font_info
                        else:
                            font_ascii = 'Times New Roman'
                            font_fareast = '宋体'
                            font_size = 12

                        # 检查文件存在
                        if not file_path or not Path(file_path).exists():
                            logger.warning(f"    ⚠️ 文件不存在: {file_path}")
                            search_start = placeholder_start + len(placeholder)
                            continue

                        # 执行替换
                        file_ext = Path(file_path).suffix.lower()
                        replace_range = doc.Range(placeholder_start, placeholder_start + len(placeholder))

                        try:
                            replace_range.Text = ' '

                            if file_ext in ['.docx', '.doc']:
                                doc_len_before = doc.Content.End
                                replace_range.InsertFile(str(Path(file_path).absolute()))
                                logger.info(f"    ✅ InsertFile: {Path(file_path).name}")
                                insert_end = doc.Content.End
                                inserted_len = insert_end - doc_len_before
                                if inserted_len > 0:
                                    insert_range = doc.Range(placeholder_start, placeholder_start + inserted_len)
                                    for para in insert_range.Paragraphs:
                                        para.Range.Font.NameAscii = font_ascii
                                        para.Range.Font.NameFarEast = font_fareast
                                        para.Range.Font.Size = font_size
                            elif file_ext == '.rtf':
                                replace_range.InsertFile(str(Path(file_path).absolute()))
                                logger.info(f"    ✅ InsertFile (RTF): {Path(file_path).name}")

                            inserted_resources.append(f"{control_title}:{placeholder}")

                        except Exception as replace_err:
                            logger.error(f"    ❌ 替换失败: {replace_err}")
                            replace_errors.append(f"{control_title}:{placeholder}")
                            # 替换失败时跳过该占位符，避免下一轮 Find 再次命中导致死循环
                            search_start = placeholder_start + len(placeholder)
                            continue

                        # 更新搜索起点（替换成功：占位符已被替换为文件内容，从其后继续）
                        search_start = placeholder_start + 1

                    except Exception as check_err:
                        logger.warning(f"    ⚠️ 检查控件失败: {check_err}")
                        search_start = placeholder_start + len(placeholder)
                        continue

                if inserted_resources:
                    logger.info(f"✅ 占位符替换完成，共 {len(inserted_resources)} 个")
                if replace_errors:
                    logger.warning(f"替换错误: {len(replace_errors)} 个")

                # 保存文档
                try:
                    doc.SaveAs(temp_output)
                    logger.info(f"✅ 第二步完成，保存文档: {Path(output_file).name}")
                except Exception as save_err:
                    logger.error(f"  ❌ 第二步保存失败: {save_err}")

                try:
                    doc.Close(False)
                    logger.info("✅ 文档已关闭")
                except Exception as close_err:
                    logger.warning(f"关闭文档时出错: {close_err}")
            else:
                logger.info("✅ 没有占位符需要替换，跳过第二步")

            # ===== 第三步：删除模板表格区域（如果指定了 replace_tag）=====
            # 遍历 generation_results，对每个表格段落检查是否需要删除模板表格
            delete_tags = []
            for result in generation_results:
                if result.get("status") != "success":
                    continue
                is_table = str(result.get("is_table", "false")).lower() == "true"
                replace_tag = result.get("replace_tag")
                if is_table and replace_tag:
                    delete_tags.append(replace_tag)

            if delete_tags:
                logger.info("=" * 70)
                logger.info(f"第三步：删除 {len(delete_tags)} 个模板表格区域")
                logger.info("=" * 70)

                for tag in delete_tags:
                    success = self.delete_template_table_region(temp_output, tag)
                    if success:
                        logger.info(f"  ✅ 已删除模板表格: {tag}")
                    else:
                        logger.warning(f"  ⚠️ 删除模板表格失败或未找到: {tag}")

                logger.info("✅ 模板表格清理完成")
            else:
                logger.info("✅ 没有需要删除的模板表格区域")

            logger.info("=" * 70)
            logger.info("✅ 插入完成")
            logger.info("=" * 70)

            return ContentInsertResult(
                success=True,
                message="内容插入成功",
                output_file=output_file,
                inserted_controls=inserted_controls,
                inserted_resources=inserted_resources,
                resource_orientations=resource_orientations  # 🆕 返回方向信息给Linux
            )

        except Exception as e:
            logger.error(f"❌ 插入失败: {e}")
            import traceback
            traceback.print_exc()
            # ✅ 出错时也要尝试关闭文档，否则文件会被锁定
            try:
                if 'doc' in dir() or 'doc' in locals():
                    doc.Close(SaveChanges=False)  # False = 不保存，直接关闭
                    logger.info("🔒 异常退出：文档已关闭（不保存）")
            except Exception as close_err:
                logger.warning(f"异常退出时关闭文档失败: {close_err}")

            return ContentInsertResult(
                success=False,
                message="内容插入失败",
                error=str(e)
            )

        finally:
            self._cleanup()

    def delete_template_table_region(self, doc_path: str, replace_tag: str) -> bool:
        """
        用 python-docx 删除模板表格区域

        删除范围：{{Table_X_Start}} 到 {{Table_X_End}} 之间的所有内容（包括标记本身）

        Args:
            doc_path: Word文档路径
            replace_tag: 要删除的标记（如 Table_1）

        Returns:
            是否删除成功
        """
        if not replace_tag:
            return False

        try:
            from docx import Document
            import copy

            logger.info(f"删除模板表格区域: {replace_tag}")

            doc = Document(doc_path)
            body = doc._element.body
            elements = list(body)

            start_tag = f"{{{{{replace_tag}_Start}}}}"
            end_tag = f"{{{{{replace_tag}_End}}}}"

            # 辅助函数：从XML元素提取文本
            def get_element_text(el) -> str:
                texts = []
                for node in el.iter():
                    tag = str(getattr(node, 'tag', '') or '')
                    if tag.endswith('t') and getattr(node, 'text', None):
                        texts.append(node.text)
                return ''.join(texts).strip()

            # 找到 Start 和 End 标记的索引
            start_idx = None
            end_idx = None

            for i, el in enumerate(elements):
                text = get_element_text(el)
                if text == start_tag:
                    start_idx = i
                elif text == end_tag and start_idx is not None:
                    end_idx = i
                    break

            if start_idx is not None and end_idx is not None:
                # 删除整个区域（从 Start 到 End，包括标记本身）
                logger.info(f"找到标记区域: index {start_idx} 到 {end_idx}")
                for el in elements[start_idx:end_idx + 1]:
                    body.remove(el)
                logger.info(f"已删除 {end_idx - start_idx + 1} 个元素")
                doc.save(doc_path)
                return True
            else:
                logger.warning(f"未找到标记区域: {replace_tag}")
                return False

        except Exception as e:
            import traceback
            traceback.print_exc()
            logger.error(f"删除模板表格区域失败: {e}")
            return False

    @staticmethod
    def test_word_control_insert(self):
        """测试Word控件内容插入 - IDE调试版本"""

        # ========== 在这里修改你的测试参数 ==========
        word_file_path = r"D:\Code\ai\4028829e9d0abef0019d23aa03bd0430.doc"  # 修改为你的Word文件路径
        control_title = "main_research_objectives"  # 修改为你的控件标题
        new_content = """队列1主要目的：评价高脂餐对健康参与者口服GFH375片的药代动力学影响。
队列2主要目的：评价艾司奥美拉唑镁肠溶片对健康参与者口服GFH375片的药代动力学影响。"""
        # 修改为你要插入的内容"

        # ==========================================

        word = None
        doc = None

        try:
            print(f"📄 打开文件: {word_file_path}")
            print(f"🎯 控件标题: {control_title}")
            print(f"📝 新内容: {new_content}")
            print("-" * 50)
            new_content = self._clean_text_for_word(new_content)

            # 启动Word
            import win32com as win32com
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False  # 显示Word窗口
            word.DisplayAlerts = 0
            # 打开文档
            doc = word.Documents.Open(
                str(word_file_path),
                ConfirmConversions=False,
                ReadOnly=False,
                AddToRecentFiles=False
            )
            print("✅ 文档打开成功")

            # 查找控件
            cc_collection = doc.SelectContentControlsByTitle(control_title)
            print(f"🔍 找到控件数量: {cc_collection.Count}")
            if cc_collection.Count == 0:
                print(f"❌ 未找到控件: {control_title}")
                return False
            for i in range(cc_collection.Count):
                control = cc_collection.Item(i + 1)
                print(f"\n--- 控件 {i + 1} ---")
                control.Range.Text = new_content

            # 查找控件
            cc_collection = doc.SelectContentControlsByTitle('project_name')
            print(f"🔍 找到控件数量: {cc_collection.Count}")
            if cc_collection.Count == 0:
                print(f"❌ 未找到控件: {control_title}")
                return False
            for i in range(cc_collection.Count):
                control = cc_collection.Item(i + 1)
                print(f"\n--- 控件 {i + 1} ---")
                control.Range.Text = 'p[pppp'


            # 保存
            doc.Save()
            print(f"\n✅ 文档已保存")
            return True

        except Exception as e:
            print(f"❌ 错误: {e}")
            return False

        finally:
            # 关闭文档
            if doc:
                try:
                    doc.Close()
                except:
                    pass

            # 退出Word
            if word:
                try:
                    word.Quit()
                except:
                    pass

if __name__ == "__main__":
    import re

    # 1. 统一换行符
    new_content = """队列1主要目的：评价高脂餐对健康参与者口服GFH375片的药代动力学影响。
    队列2主要目的：评价艾司奥美拉唑镁肠溶片对健康参与者口服GFH375片的药代动力学影响。"""
    text = new_content.replace('\r\n', '\r')
    print(text)
    # import pythoncom
    # pythoncom.CoInitialize()
    # inserter = WordControlContentInserter()  # 创建实例
    # inserter.test_word_control_insert()


