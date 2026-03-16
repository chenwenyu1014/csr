# -*- coding: utf-8 -*-
from __future__ import annotations

import logging
import math
import os
import re
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from enum import Enum

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

logger = logging.getLogger(__name__)


# ContentType enum
class ContentType(Enum):
    MARKDOWN = "markdown"
    STRUCTURED = "structured"


def run(excel_path: Path | str, work_dir: Path | str) -> Dict[str, Any]:
    excel_path = Path(excel_path)
    work_dir = Path(work_dir)

    sheets_dir = work_dir / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)

    # 1) Split sheets to individual xlsx files
    regions = split_sheets(excel_path, sheets_dir)

    # 2) Convert each sheet to Markdown (manually iterate over split sheets)
    md_dir = sheets_dir / "markdown"
    md_dir.mkdir(parents=True, exist_ok=True)
    md_regions: List[Dict[str, Any]] = []
    
    for sheet_info in regions:
        sheet_file = Path(sheet_info['file_path'])
        sheet_name = sheet_info.get('sheet_name', 'Sheet')
        if sheet_file.exists():
            try:
                md_text = convert_excel_sheets_to_markdown(
                    excel_path=sheet_file,
                    output_dir=md_dir,
                    header_rows=1,
                    placeholder="...",
                    fill_mode="dots",
                    trim=True,
                    percent_dp=2,
                    enable_percent=True
                )
                # md_text is a dict with 'success', 'sheets', etc.
                if md_text.get('success'):
                    # 对于单个sheet文件，应该生成一个MD文件
                    # MD文件名通常为: {sheet_name}.md
                    md_file_path = md_dir / f"{sheet_name}.md"
                    if md_file_path.exists():
                        md_regions.append({
                            'sheet_name': sheet_name,
                            'markdown_file': str(md_file_path)
                        })
                    else:
                        # 如果使用sheets返回值
                        for sheet in md_text.get('sheets', []):
                            md_file = sheet.get('markdown_file')
                            if md_file:
                                md_path = Path(md_file) if Path(md_file).is_absolute() else md_dir / md_file
                                md_regions.append({
                                    'sheet_name': sheet_name,
                                    'markdown_file': str(md_path)
                                })
            except Exception as e:
                import logging
                logging.warning(f"Failed to convert sheet {sheet_name} to markdown: {e}")

    # 3) Normalize md_regions to include sheet_name
    md_map = {}
    for r in md_regions:
        # r: { 'sheet_name', 'markdown_file', 'rows', 'cols' }
        md_map[r.get('sheet_name') or ''] = r

    # pair sheet_name
    md_regions_norm: List[Dict[str, Any]] = []
    for r in regions:
        sname = r.get('sheet_name', '')
        md_info = md_map.get(sname)
        if md_info:
            md_regions_norm.append(md_info)

    return {
        'content': '',
        'content_type': ContentType.STRUCTURED,
        'text': '',
        'excel_regions': regions,
        'excel_markdown_regions': md_regions_norm,
        'metadata': {
            'conversion_method': 'excel_pipeline',
            'steps': ['excel_split_sheets', 'excel_sheets_to_markdown'],
            'sheets_total': len(regions),
            'excel_sheets': regions,
            'excel_markdown_regions': md_regions_norm,
            'out_dir': str(sheets_dir.resolve()),
        }
    }



#  Excel 划分sheet 工具
def _com_enabled() -> bool:
    v = str(os.getenv('EXCEL_COM_ENABLED', '0')).strip().lower()
    return v in ('1', 'true', 'yes', 'y', 'on')


def split_sheets(excel_path: Path | str, out_dir: Path | str) -> List[Dict[str, Any]]:
    excel_path = Path(excel_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    regions: List[Dict[str, Any]] = []
    logger.info(f"split_sheets start | excel={excel_path} | out_dir={out_dir} | com_enabled={_com_enabled()}")

    if _com_enabled():
        try:
            import pythoncom  # type: ignore
            import win32com.client as win32  # type: ignore
            pythoncom.CoInitialize()
            from utils.windows_com import safe_dispatch
            excel = safe_dispatch("Excel.Application", use_ex=True, logger=logger)
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(excel_path.resolve()))
            try:
                logger.info(f"COM split open ok | sheets={wb.Worksheets.Count}")
                for i, ws in enumerate(wb.Worksheets, start=1):  # type: ignore
                    name = ws.Name
                    safe_name = _safe_sheet_name(name)
                    sheet_file = out_dir / f"{excel_path.stem}_{safe_name}.xlsx"
                    # Save each sheet as its own workbook
                    ws.Copy()
                    new_wb = excel.ActiveWorkbook
                    try:
                        new_wb.SaveAs(str(sheet_file.resolve()))
                        logger.debug(f"COM saved sheet | idx={i} | name={name} | path={sheet_file}")
                    finally:
                        new_wb.Close(SaveChanges=False)
                    regions.append({'sheet_name': name, 'file_path': str(sheet_file)})
            finally:
                wb.Close(SaveChanges=False)
                excel.Quit()
                pythoncom.CoUninitialize()
            logger.info(f"split_sheets done via COM | regions={len(regions)}")
            return regions
        except Exception as e:
            logger.warning(f"Excel COM 拆分失败，回退 openpyxl: {e}")
            # fall back

    # Fallback: openpyxl
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(str(excel_path), data_only=False)
        try:
            logger.info(f"openpyxl split open ok | sheets={len(wb.worksheets)}")
            for ws in wb.worksheets:
                name = ws.title
                safe_name = _safe_sheet_name(name)
                sheet_file = out_dir / f"{excel_path.stem}_{safe_name}.xlsx"
                from openpyxl import Workbook  # type: ignore
                nwb = Workbook()
                nws = nwb.active
                nws.title = name
                # copy cell values and basic merges
                for row in ws.iter_rows(values_only=True):
                    nws.append(list(row))
                # Copy merged cells
                try:
                    for merged in ws.merged_cells.ranges:
                        nws.merge_cells(str(merged))
                except Exception:
                    pass
                nwb.save(str(sheet_file))
                logger.debug(f"openpyxl saved sheet | name={name} | path={sheet_file}")
                regions.append({'sheet_name': name, 'file_path': str(sheet_file)})
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception as e:
        logger.error(f"openpyxl 拆分失败: {e}")
        raise

    logger.info(f"split_sheets done via openpyxl | regions={len(regions)}")
    return regions


def _safe_sheet_name(name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = ''.join('_' if ch in invalid else ch for ch in name).strip()
    return cleaned[:31] or 'Sheet'


# Excel 转 Markdown 工具部分
def parse_a1_range(a1: str) -> Tuple[int, int, int, int]:
    """解析 A1 区域字符串为 (min_row, min_col, max_row, max_col)"""
    if ":" not in a1:
        r1, c1 = coordinate_to_tuple(a1)
        return r1, c1, r1, c1
    left, right = a1.split(":", 1)
    r1, c1 = coordinate_to_tuple(left)
    r2, c2 = coordinate_to_tuple(right)
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def detect_used_range(ws) -> Tuple[int, int, int, int]:
    """检测工作表的已用区域"""
    min_row, min_col = 10**9, 10**9
    max_row, max_col = 0, 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v not in (None, ""):
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    if max_row == 0:
        return 1, 1, 1, 1
    return min_row, min_col, max_row, max_col


def build_grid(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> List[List]:
    """构建单元格对象网格"""
    h = max_row - min_row + 1
    w = max_col - min_col + 1
    grid = [[None for _ in range(w)] for _ in range(h)]
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            grid[r - min_row][c - min_col] = ws.cell(row=r, column=c)
    return grid


def expand_merged_cells(ws, values: List[List], base_r: int, base_c: int,
                       placeholder: str, fill_mode: str):
    """
    展开合并单元格
    fill_mode='dots': 左上角保留原值，其余填充占位符
    fill_mode='copy': 所有拆分格复制左上角原值
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        min_row, min_col, max_row, max_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        tl_val = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                rr = r - base_r
                cc = c - base_c
                if rr < 0 or cc < 0 or rr >= len(values) or cc >= len(values[0]):
                    continue
                if r == min_row and c == min_col:
                    values[rr][cc] = tl_val
                else:
                    if fill_mode == "copy":
                        values[rr][cc] = tl_val
                    else:
                        values[rr][cc] = placeholder


def strip_empty_edges(values: List[List]) -> List[List]:
    """去掉首尾全空行/列"""
    def row_empty(row):
        return all((v is None or (isinstance(v, str) and v.strip() == "")) for v in row)

    def col_empty(vals, ci):
        for r in vals:
            v = r[ci]
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                return False
        return True

    # 去掉尾部空行
    while len(values) > 1 and row_empty(values[-1]):
        values.pop()
    # 去掉顶部空行
    while len(values) > 1 and row_empty(values[0]):
        values.pop(0)

    if not values:
        return values

    # 去掉右侧空列
    while len(values[0]) > 1 and col_empty(values, len(values[0]) - 1):
        for r in values:
            r.pop()
    # 去掉左侧空列
    while len(values[0]) > 1 and col_empty(values, 0):
        for r in values:
            r.pop(0)

    return values


def sanitize_text(s: str) -> str:
    """转义管道符，规范化空白"""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = s.replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def fmt_value(cell, placeholder: str, percent_dp: int, enable_percent: bool) -> str:
    """格式化单元格值"""
    v = cell.value
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return placeholder

    # 百分比格式化
    if enable_percent:
        try:
            nf = (cell.number_format or "").lower()
        except Exception:
            nf = ""
        if "%" in nf and isinstance(v, (int, float)):
            num = float(v) * 100.0
            q = f"{{:.{percent_dp}f}}".format(num)
            q = q.rstrip("0").rstrip(".") if percent_dp > 0 else q
            return q

    # 数值处理：避免科学计数法
    if isinstance(v, float):
        if math.isfinite(v) and abs(v - round(v)) < 1e-12:
            return str(int(round(v)))
        s = f"{v:.6f}".rstrip("0").rstrip(".")
        return s or "0"

    return str(v)


def _is_single_cell_row(row: List[str], placeholder: str = "...") -> bool:
    """
    判断是否为单单元格行（整行只有第一个单元格有内容，其余全是占位符）
    """
    if not row:
        return False

    first_cell = str(row[0]).strip() if row[0] else ""
    if not first_cell or first_cell == placeholder:
        return False

    # 检查其余单元格是否全是占位符
    for cell in row[1:]:
        cell_val = str(cell).strip() if cell else ""
        if cell_val and cell_val != placeholder:
            return False

    return True


def _is_footer_content(text: str) -> bool:
    """
    判断是否为页脚内容（需要过滤掉的内容）

    页脚特征：
    - 以 "-" 或 "- " 开头（如 "- 数据来源：xxx"）
    - 包含 "路径：" 或 "路径:"
    - 包含 "数据来源" 或 "本表汇总"
    - 包含 ".SAS" 或 ".sas"（SAS程序路径）
    - 包含 "Final" + 日期格式
    """
    import re

    text = text.strip()

    # 以 "-" 开头的说明性文本
    if text.startswith("-") or text.startswith("- "):
        return True

    # 路径信息
    if "路径：" in text or "路径:" in text:
        return True

    # 数据来源说明
    if "数据来源" in text:
        return True

    # 本表汇总说明
    if "本表汇总" in text:
        return True

    # SAS程序路径
    if ".SAS" in text or ".sas" in text:
        return True

    # Final + 日期格式（如 "Final 2025-01-07 12:36"）
    if re.search(r'Final\s+\d{4}-\d{2}-\d{2}', text, re.IGNORECASE):
        return True

    return False


def _is_title_row(row: List[str], placeholder: str = "...") -> bool:
    """
    判断是否为真正的标题行（整行合并单元格，且不是页脚内容）

    标题行特征：
    - 第一个单元格有实际内容
    - 其余单元格全是占位符（"..."）或空白
    - 不是页脚内容

    例如：["表 14.3-13.1.1 单次给药-注射部位观察", "...", "...", "..."]
    """
    if not _is_single_cell_row(row, placeholder):
        return False

    first_cell = str(row[0]).strip()

    # 过滤掉页脚内容
    if _is_footer_content(first_cell):
        return False

    return True


def _is_footer_row(row: List[str], placeholder: str = "...") -> bool:
    """
    判断是否为页脚行（需要过滤掉的行）
    """
    if not _is_single_cell_row(row, placeholder):
        return False

    first_cell = str(row[0]).strip()
    return _is_footer_content(first_cell)


def to_markdown(
    values: List[List[str]],
    header_rows: int = 1,
    placeholder: str = "...",
    sheet_title: Optional[str] = None,
) -> str:
    """
    转换为 Markdown 表格

    增强：
    1. 自动识别标题行（整行合并单元格），将其作为独立标题
    2. 自动过滤页脚行（数据来源、路径等说明性文本）

    标题行特征：整行只有第一个单元格有内容，其余全是占位符
    例如：["表 14.3-13.1.1 单次给药-注射部位观察", "...", "...", "..."]

    页脚行特征：以 "-" 开头、包含"路径："、"数据来源"等
    """
    if not values:
        return ""

    lines = []
    title_lines = []  # 收集标题行
    data_rows = []    # 收集数据行

    # 分离标题行、数据行、过滤页脚行
    for row in values:
        if _is_footer_row(row, placeholder):
            # 这是页脚行，跳过（不输出）
            continue
        elif _is_title_row(row, placeholder):
            # 这是标题行，提取第一个单元格作为标题（不加 ## 前缀）
            title_text = str(row[0]).strip()
            title_lines.append(title_text)
        else:
            data_rows.append(row)

    # 添加标题（作为普通文本，不加 Markdown 标题格式）
    # 始终输出 sheet 名到正文开头（避免只当作 sheet 名而正文缺失）
    # 若表内已存在标题行，则避免重复（忽略首尾空格比较）
    titles_to_output: List[str] = []
    if sheet_title and sheet_title.strip():
        titles_to_output.append(sheet_title.strip())
    for t in title_lines:
        if t.strip() not in titles_to_output:
            titles_to_output.append(t.strip())
    if titles_to_output:
        lines.extend(titles_to_output)
        lines.append("")  # 空行分隔

    # 生成表格
    if data_rows:
        # 使用数据行的列数，而不是原始values的列数
        ncols = len(data_rows[0]) if data_rows else 0

        for ri, row in enumerate(data_rows):
            # 确保行的列数与ncols一致
            padded_row = row + [placeholder] * (ncols - len(row)) if len(row) < ncols else row[:ncols]
            line = "| " + " | ".join(padded_row) + " |"
            lines.append(line)
            if ri == header_rows - 1:
                sep = "| " + " | ".join(["---"] * ncols) + " |"
                lines.append(sep)

        if header_rows <= 0 and len(data_rows) > 0:
            sep = "| " + " | ".join(["---"] * ncols) + " |"
            lines.insert(len(title_lines) + 1 if title_lines else 1, sep)

    return "\n".join(lines)


def excel_sheet_to_markdown(
    excel_path: Path,
    sheet_name: Optional[str] = None,
    a1_range: Optional[str] = None,
    header_rows: int = 1,
    placeholder: str = "...",
    fill_mode: str = "dots",
    trim: bool = True,
    percent_dp: int = 1,
    enable_percent: bool = True
) -> str:
    """
    将 Excel 工作表转换为 Markdown 表格

    Args:
        excel_path: Excel 文件路径
        sheet_name: 工作表名称（None 则取 active）
        a1_range: 区域范围（如 "A1:H20"，None 则自动检测）
        header_rows: 表头行数
        placeholder: 空白填充值
        fill_mode: 合并单元格填充模式（'dots' 或 'copy'）
        trim: 是否裁剪首尾空行/列
        percent_dp: 百分比小数位
        enable_percent: 是否启用百分比格式化

    Returns:
        Markdown 表格文本
    """
    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 确定区域
    if a1_range:
        min_row, min_col, max_row, max_col = parse_a1_range(a1_range)
    else:
        min_row, min_col, max_row, max_col = detect_used_range(ws)

    # 构建网格
    cell_grid = build_grid(ws, min_row, min_col, max_row, max_col)

    # 提取初始值
    h = len(cell_grid)
    w = len(cell_grid[0]) if h > 0 else 0
    values = [[cell_grid[r][c].value if cell_grid[r][c] is not None else None
               for c in range(w)] for r in range(h)]

    # 展开合并单元格
    expand_merged_cells(ws, values, base_r=min_row, base_c=min_col,
                       placeholder=placeholder, fill_mode=fill_mode)

    # 保留非合并单元格原值
    for r in range(h):
        for c in range(w):
            if values[r][c] is None:
                values[r][c] = cell_grid[r][c].value if cell_grid[r][c] else None

    # 填充占位符
    for r in range(h):
        for c in range(w):
            v = values[r][c]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                values[r][c] = placeholder

    # 可选裁剪
    if trim:
        values = strip_empty_edges(values)

    # 格式化为字符串
    rendered: List[List[str]] = []
    for r, row in enumerate(values):
        row_out = []
        for c, v in enumerate(row):
            cell = cell_grid[r][c] if r < len(cell_grid) and c < len(cell_grid[0]) else None
            if cell is None:
                text = placeholder if (v is None or str(v).strip() == "") else str(v)
            else:
                text = fmt_value(cell, placeholder=placeholder, percent_dp=percent_dp,
                               enable_percent=enable_percent)
            row_out.append(sanitize_text(text))
        rendered.append(row_out)

    # 生成 Markdown
    md = to_markdown(
        rendered,
        header_rows=header_rows,
        placeholder=placeholder,
        sheet_title=ws.title if ws else None,  # 兜底使用 sheet 名
    )

    wb.close()
    return md


def convert_excel_sheets_to_markdown(
    excel_path: Path,
    output_dir: Path,
    header_rows: int = 1,
    placeholder: str = "...",
    fill_mode: str = "dots",
    trim: bool = True,
    percent_dp: int = 1,
    enable_percent: bool = True
) -> Dict[str, Any]:
    """
    将 Excel 所有工作表转换为 Markdown 文件

    Returns:
        {
            'success': bool,
            'sheets': [
                {
                    'sheet_name': str,
                    'markdown_file': str (相对路径),
                    'rows': int,
                    'cols': int
                }
            ],
            'error': str (如果有错误)
        }
    """
    try:
        output_dir.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(str(excel_path), data_only=True)
        sheets_info = []

        for sheet_name in wb.sheetnames:
            try:
                md = excel_sheet_to_markdown(
                    excel_path,
                    sheet_name=sheet_name,
                    header_rows=header_rows,
                    placeholder=placeholder,
                    fill_mode=fill_mode,
                    trim=trim,
                    percent_dp=percent_dp,
                    enable_percent=enable_percent
                )

                # 保存为 Markdown 文件
                safe_sheet_name = re.sub(r'[<>:"/\\|?*]', '_', sheet_name)
                md_file = output_dir / f"{safe_sheet_name}.md"
                md_file.write_text(md, encoding='utf-8')

                # 统计行列数
                lines = md.split('\n')
                rows = len([l for l in lines if l.startswith('|')])
                cols = len([l for l in lines if l.startswith('|')][0].split('|')) - 2 if rows > 0 else 0

                sheets_info.append({
                    'sheet_name': sheet_name,
                    'markdown_file': f"{safe_sheet_name}.md",
                    'rows': rows,
                    'cols': cols
                })
            except Exception as e:
                sheets_info.append({
                    'sheet_name': sheet_name,
                    'error': str(e)
                })

        wb.close()

        return {
            'success': True,
            'sheets': sheets_info,
            'total_sheets': len(sheets_info)
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }
